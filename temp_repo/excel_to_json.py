# -*- coding: utf-8 -*-
"""
사주명리_완전정리_DB.xlsx → saju_db.json 변환기
────────────────────────────────────────────────────────────
GitHub Actions 자동화용. DB 수정 후 재변환.

사용법:
  python excel_to_json.py                     # 기본 경로 사용
  python excel_to_json.py <xlsx_path>         # 특정 xlsx 경로 지정
  python excel_to_json.py <xlsx_path> <json_out>
                                              # 출력 경로까지 지정

동작:
  1. 엑셀의 각 시트를 파싱하여 JSON 구조로 변환
  2. 시트18 기반으로 시트19(일주별 직업군 매칭)를 자동 재생성
  3. 시트18 + 시트17 기반으로 시트20(월지별 직업군 매칭)을 자동 생성
     → 시트18 편집만 하면 시트19·20은 매 실행마다 동기화됨
  4. saju_db.json 하나의 파일로 통합 출력

의존성:
  pip install pandas openpyxl

JSON 구조:
  {
    "dna": {...},                   # 시트12 — 60갑자 common/male/female DNA
    "gender_traits": {...},         # 시트10 — 천간×성별 특성
    "ilju_unique": {...},           # 시트11 — 60일주 고유 성격/외형
    "sipsung_jobs": {...},          # 시트6 — 십성별 직업방향
    "job_visual": {...},            # 시트14 — 직업별 비주얼 DB
    "body_pct": {...},              # 시트15 — 60일주 × 7카테고리 신체 퍼센트
    "ilju_combo": {...},            # 시트16 — 조합 키워드/격국/설명
    "month_correction": {...},      # 시트17 — 12월지 보정
    "branch_harmony": {...},        # 시트17 하단 — 지지 합충 관계표
    "job_categories_100": [...],    # 시트18 — 100 직업군(분리 태그 포함)
    "job_exclude_option": {...},    # 시트18 JC000 행
    "ilju_to_jobs": {...},          # 시트19 자동 재생성 — 일주 역인덱스
    "branch_to_jobs": {...},        # 시트20 자동 재생성 — 월지 역인덱스
    "meta": {...}                   # 빌드 메타데이터
  }

병합 가이드 (index.html 사용법):
  일주 A × 월지 B 선택 시:
    primary   = ilju_to_jobs[A] ∩ branch_to_jobs[B]   # 강력 추천
    by_ilju   = ilju_to_jobs[A] − branch_to_jobs[B]   # 일주 기반
    by_branch = branch_to_jobs[B] − ilju_to_jobs[A]   # 월지 기반 (신규 확장)
    union     = ilju_to_jobs[A] ∪ branch_to_jobs[B]   # 전체 선택 가능
"""
import os
import sys
import json
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl


# ─────────────────────────────────────────────────────────────
# 경로 처리
# ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, '사주명리_완전정리_DB.xlsx')
DEFAULT_JSON = os.path.join(SCRIPT_DIR, 'saju_db.json')


# ─────────────────────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────────────────────
def cell(df, r, c):
    """pandas DataFrame의 (r, c) 셀을 안전하게 문자열로"""
    if r >= len(df) or c >= len(df.columns):
        return ''
    v = df.iloc[r, c]
    return str(v).strip() if pd.notna(v) and str(v).strip() else ''


def split_list(s, seps=(',', '·', '/')):
    """문자열을 여러 구분자로 잘라 공백 제거 후 반환"""
    if not s:
        return []
    tokens = [s]
    for sep in seps:
        new = []
        for t in tokens:
            new.extend(t.split(sep))
        tokens = new
    return [t.strip() for t in tokens if t.strip() and t.strip() != '—']


# ─────────────────────────────────────────────────────────────
# 시트12 — 일주 캐릭터 DNA
# ─────────────────────────────────────────────────────────────
COL12 = {
    4: 'A1_골격', 5: 'A2_체형',
    6: 'B1_얼굴형', 7: 'B2_눈', 8: 'B3_이면눈매',
    9: 'B4_눈썹', 10: 'B5_코', 11: 'B6_입', 12: 'B7_얼굴음영',
    13: 'C1_피부톤', 14: 'C2_오행변이',
    15: 'D1_헤어', 16: 'D2_머리물리', 17: 'D3_수염',
    18: 'E1_의복핏', 19: 'E2_상의', 20: 'E3_하의', 21: 'E4_풀바디', 22: 'E5_신발',
    23: 'F1_머리악세', 24: 'F2_몸악세', 25: 'F3_소울웨폰', 26: 'F4_장신구공명',
    27: 'F5_인장문양', 28: 'F6_특수요소',
    29: 'G1_아우라', 30: 'G2_이펙트컬러', 31: 'G3_그림자', 32: 'G4_발걸음',
    33: 'G5_동작잔상', 34: 'G6_지장간',
    35: 'H1_표정온도', 36: 'H2_분위기', 37: 'H3_제스처', 38: 'H4_음성톤',
    39: 'H5_감정중력', 40: 'H6_각성트리거', 41: 'H7_나이대종족',
}


def extract_dna(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트12_일주_캐릭터DNA', header=None)
    dna = {}
    cur = None
    for i in range(3, len(df)):
        ilju = cell(df, i, 1)
        if ilju:
            cur = ilju
        if not cur:
            continue
        g = {'共': 'common', '♂': 'male', '♀': 'female'}.get(cell(df, i, 3))
        if not g:
            continue
        if cur not in dna:
            dna[cur] = {}
        rd = {}
        for ci, cn in COL12.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        dna[cur][g] = rd
    return dna


# ─────────────────────────────────────────────────────────────
# 시트10 — 천간×성별 특성
# ─────────────────────────────────────────────────────────────
def extract_gender(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트10_성별', header=None)
    gt = {}
    cols = {
        3: '성격분기', 4: '외형분기', 5: '체형분기', 6: '피부인상분기',
        7: '눈빛표정분기', 8: '분위기태그', 9: '스타일분기',
        10: '직업방향', 11: '관계패턴', 12: '프롬프트외형태그',
    }
    for i in range(2, len(df)):
        cg = cell(df, i, 1)
        gd = cell(df, i, 2)
        if not cg or not gd:
            continue
        hj = cg.split('(')[1].split(')')[0] if '(' in cg else cg
        key = f"{hj}_{gd}"
        rd = {}
        for ci, cn in cols.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        gt[key] = rd
    return gt


# ─────────────────────────────────────────────────────────────
# 시트11 — 60일주 고유 특성
# ─────────────────────────────────────────────────────────────
def extract_ilju(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트11_일주', header=None)
    iu = {}
    cols = {
        3: '천간', 4: '지지', 5: '납음', 6: '납음해석', 7: '십이운성',
        8: '일지십성', 9: '생극관계', 10: '고유성격', 11: '고유외형', 12: '고유이미지',
    }
    for i in range(2, len(df)):
        ilju = cell(df, i, 2)
        if not ilju:
            continue
        rd = {}
        for ci, cn in cols.items():
            v = cell(df, i, ci)
            if v:
                rd[cn] = v
        iu[ilju] = rd
    return iu


# ─────────────────────────────────────────────────────────────
# 시트6 — 십성별 직업방향
# ─────────────────────────────────────────────────────────────
def extract_sipsung_jobs(xlsx):
    df = pd.read_excel(xlsx, sheet_name='시트6_지지십성', header=None)
    sj = {}
    for i in range(2, len(df)):
        name = cell(df, i, 1)
        job = cell(df, i, 6)
        if name and job:
            sj[name] = job
    return sj


# ─────────────────────────────────────────────────────────────
# 시트14 — 직업별 비주얼
# ─────────────────────────────────────────────────────────────
def extract_job_visual(xlsx):
    try:
        df = pd.read_excel(xlsx, sheet_name='시트14_직업비주얼', header=None)
    except Exception:
        return {}
    jv = {}
    for i in range(2, len(df)):
        name = cell(df, i, 1)
        desc = cell(df, i, 2)
        sd = cell(df, i, 3)
        if name:
            entry = {}
            if desc:
                entry['desc'] = desc
            if sd:
                entry['sd'] = sd
            jv[name] = entry
    return jv


# ─────────────────────────────────────────────────────────────
# 시트15 — 60일주 신체 퍼센트 매트릭스
# ─────────────────────────────────────────────────────────────
# 23개 퍼센트 컬럼 (C~Y)을 7개 카테고리로 묶음
BODY_LAYOUT = [
    ('키',       4,  ['단신', '중간', '장신', '매우장신']),
    ('체형',     4,  ['마른', '보통', '근육', '풍채']),
    ('나잇대',   3,  ['동안', '또래', '성숙']),
    ('얼굴',     3,  ['작은편', '보통', '큰편']),
    ('이목구비', 3,  ['섬세', '보통', '뚜렷']),
    ('피부',     3,  ['백옥', '건강', '구릿빛']),
    ('모발',     3,  ['가늘', '보통', '굵은']),
]


def extract_body_pct(xlsx):
    """
    시트15 구조:
      - 헤더 2개 행 (병합 대분류 + 하위 세부)
      - 4번째 행부터 데이터 (A: 번호, B: 일주, C~: 퍼센트)
    """
    try:
        df = pd.read_excel(xlsx, sheet_name='시트15_일주신체분포', header=None)
    except Exception:
        return {}

    body = {}
    for i in range(3, len(df)):
        ilju = cell(df, i, 1)
        if not ilju:
            continue
        row_data = {}
        col_idx = 2  # C열부터 시작
        for cat_name, n, sublabels in BODY_LAYOUT:
            pcts = []
            for k in range(n):
                v = df.iloc[i, col_idx + k] if col_idx + k < len(df.columns) else 0
                try:
                    pcts.append(int(v) if pd.notna(v) else 0)
                except (ValueError, TypeError):
                    pcts.append(0)
            row_data[cat_name] = pcts
            col_idx += n
        body[ilju] = row_data
    return body


# ─────────────────────────────────────────────────────────────
# 시트16 — 60일주 조합 키워드
# ─────────────────────────────────────────────────────────────
def extract_ilju_combo(xlsx):
    try:
        df = pd.read_excel(xlsx, sheet_name='시트16_일주조합키워드', header=None)
    except Exception:
        return {}

    combo = {}
    for i in range(2, len(df)):
        ilju = cell(df, i, 1)
        if not ilju:
            continue
        combo[ilju] = {
            'kw_ko':      split_list(cell(df, i, 2), seps=(' · ',)),
            'kw_en':      [x.strip() for x in cell(df, i, 3).split(',') if x.strip()],
            '격국':       cell(df, i, 4),
            '특수격':     cell(df, i, 5),
            '오행강약':   cell(df, i, 6),
            '설명':       cell(df, i, 7),
        }
    return combo


# ─────────────────────────────────────────────────────────────
# 시트17 — 월지 보정 + 지지 합충 관계표
# ─────────────────────────────────────────────────────────────
def extract_month_correction(xlsx):
    """월지 보정표(상단 12행) + 합충 관계표(하단)를 분리 반환"""
    try:
        df = pd.read_excel(xlsx, sheet_name='시트17_월지보정', header=None)
    except Exception:
        return {}, {}

    month = {}
    # 상단 12행: 월지 보정
    for i in range(2, len(df)):
        br = cell(df, i, 0)
        # 지지 1글자 + 유효 지지만 허용
        if len(br) != 1 or br not in '寅卯辰巳午未申酉戌亥子丑':
            continue
        if len(month) >= 12:
            break
        month[br] = {
            '계절':       cell(df, i, 1),
            '오행':       cell(df, i, 2),
            '조후':       cell(df, i, 3),
            '장간':       cell(df, i, 4),
            '월령십성':   cell(df, i, 5),
            '키워드':     split_list(cell(df, i, 6), seps=(' · ',)),
            '설명':       cell(df, i, 7),
        }

    # 하단 합충 관계표: 2글자 지지 조합
    harmony = {}
    for i in range(2, len(df)):
        a = cell(df, i, 0)
        b = cell(df, i, 1)
        rel = cell(df, i, 2)
        if (len(a) == 1 and len(b) == 1 and a != b
                and a in '寅卯辰巳午未申酉戌亥子丑'
                and b in '寅卯辰巳午未申酉戌亥子丑'
                and rel):
            harmony[f'{a}_{b}'] = rel

    return month, harmony


# ─────────────────────────────────────────────────────────────
# 시트18 — 직업군 100 (분리 태그 포함)
# ─────────────────────────────────────────────────────────────
def extract_jobs_100(xlsx):
    """
    시트18 구조 (10개 컬럼):
      A 코드 / B 직업군명 / C 대분류 / D 키워드 / E 오행 태그 /
      F 십성 태그 / G 기질 태그 / H 특수격 태그 / I 매칭 일주 수 /
      J 매칭 일주 목록
    """
    try:
        df = pd.read_excel(xlsx, sheet_name='시트18_직업군100', header=None)
    except Exception:
        return [], None

    jobs = []
    jc000 = None
    for i in range(2, len(df)):
        code = cell(df, i, 0)
        if not code or not code.startswith('JC'):
            continue

        matching = [
            p.strip() for p in cell(df, i, 9).split(',')
            if p.strip() and p.strip() not in ('—', '전체 60일주 선택 가능')
        ]
        keywords = split_list(cell(df, i, 3), seps=(' · ',))

        entry = {
            'id':            code,
            'name':          cell(df, i, 1),
            'category':      cell(df, i, 2),
            'keywords':      keywords,
            'oheng':         cell(df, i, 4),
            'sipsung':       cell(df, i, 5),
            'gijil':         cell(df, i, 6),
            'teukgyeok':     cell(df, i, 7),
            'matching_ilju': matching,
        }

        if code == 'JC000':
            jc000 = entry
        else:
            jobs.append(entry)
    return jobs, jc000


# ─────────────────────────────────────────────────────────────
# 시트19 자동 재생성
#   시트18의 매칭 일주 목록을 스캔하여 역인덱스 계산
#   → 엑셀 시트19에 쓰기 + 반환값은 JSON용
# ─────────────────────────────────────────────────────────────
def refresh_sheet19(xlsx, jobs):
    """시트19 자동 갱신. 반환: {일주: [JC코드, ...]}"""
    ilju_to_jobs = defaultdict(list)
    for j in jobs:
        for p in j['matching_ilju']:
            ilju_to_jobs[p].append(j['id'])
    ilju_to_jobs = dict(ilju_to_jobs)

    # 엑셀 시트19에 값 기입 (openpyxl)
    try:
        wb = openpyxl.load_workbook(xlsx)
        if '시트19_일주별직업군매칭' in wb.sheetnames:
            ws = wb['시트19_일주별직업군매칭']
            # 헤더 3행 아래 데이터는 r=4부터 (60일주)
            for r in range(4, 64):
                ilju = ws.cell(row=r, column=2).value
                if not ilju:
                    continue
                codes = ilju_to_jobs.get(ilju, [])
                ws.cell(row=r, column=3).value = len(codes)
                ws.cell(row=r, column=4).value = ', '.join(codes) if codes else ''
            wb.save(xlsx)
    except Exception as e:
        # 엑셀 쓰기 실패해도 JSON은 정상 반환 (CI 환경 등에서 파일 잠김 대비)
        print(f'  WARN: 시트19 엑셀 쓰기 실패 ({e}). JSON은 정상 생성됨.')

    return ilju_to_jobs


# ─────────────────────────────────────────────────────────────
# 월지 ↔ 직업군 매칭 로직 (시트20 자동 생성용)
# ─────────────────────────────────────────────────────────────
# 월지별 (오행 / 본기 천간 / 계절 키워드)
BRANCH_META = {
    '寅': {'오행': '木', '본기': '甲', '계절': ['초봄', '활동', '개척', '생기']},
    '卯': {'오행': '木', '본기': '乙', '계절': ['중봄', '섬세', '미', '도화', '창의']},
    '辰': {'오행': '土', '본기': '戊', '계절': ['늦봄', '저장', '수고', '전략']},
    '巳': {'오행': '火', '본기': '丙', '계절': ['초여름', '추진', '재성', '확장']},
    '午': {'오행': '火', '본기': '丁', '계절': ['중여름', '스타', '카리스마', '양인', '화려']},
    '未': {'오행': '土', '본기': '己', '계절': ['늦여름', '화고', '온화', '결실']},
    '申': {'오행': '金', '본기': '庚', '계절': ['초가을', '냉철', '결단', '정의', '기술']},
    '酉': {'오행': '金', '본기': '辛', '계절': ['중가을', '정교', '완성', '미', '도화']},
    '戌': {'오행': '土', '본기': '戊', '계절': ['늦가을', '화고', '지혜', '신비']},
    '亥': {'오행': '水', '본기': '壬', '계절': ['초겨울', '지혜', '학문', '내면']},
    '子': {'오행': '水', '본기': '癸', '계절': ['중겨울', '양인', '지혜극치', '집중']},
    '丑': {'오행': '土', '본기': '己', '계절': ['늦겨울', '금고', '인내', '축적']},
}

_STEM_OHENG = {
    '甲': '木', '乙': '木', '丙': '火', '丁': '火',
    '戊': '土', '己': '土', '庚': '金', '辛': '金',
    '壬': '水', '癸': '水',
}

# 월지별 특수격 가중 (괴강·양인·건록·도화 등)
_BRANCH_SPECIAL = {
    '午': ['양인'], '子': ['양인'],
    '寅': ['건록'], '卯': ['건록', '도화'],
    '巳': ['건록'], '申': ['건록'],
    '酉': ['건록', '도화'], '亥': ['건록'],
    '辰': ['괴강'], '戌': ['괴강'],
}


def _score_branch_job(branch, job):
    """
    월지와 직업군의 매칭 점수 계산.
    신호 가중치:
      1. 월지 오행 ∈ 직업군 오행 태그 → +2
      2. 본기 천간 오행이 직업군 오행에 포함(주오행과 다를 때) → +1
      3. 월지 특수격 ∈ 직업군 특수격 태그 → +1
      4. 계절 키워드 ∈ 직업군 키워드·기질 → +1
    """
    if branch not in BRANCH_META:
        return 0
    meta = BRANCH_META[branch]
    score = 0

    oheng_tag = job.get('oheng') or ''
    if meta['오행'] in oheng_tag:
        score += 2

    bonki_oheng = _STEM_OHENG.get(meta['본기'], '')
    if bonki_oheng and bonki_oheng != meta['오행'] and bonki_oheng in oheng_tag:
        score += 1

    teuk = job.get('teukgyeok') or ''
    if teuk and teuk != '—':
        for s in _BRANCH_SPECIAL.get(branch, []):
            if s in teuk:
                score += 1
                break

    job_tokens = (job.get('keywords') or []) + (job.get('gijil') or '').split('·')
    joined = ' '.join(job_tokens)
    for season_kw in meta['계절']:
        if season_kw in joined:
            score += 1
            break

    return score


def build_branch_to_jobs(jobs, threshold=2):
    """
    12월지 각각에 대해 매칭 직업군 코드 리스트(점수 내림차순) 반환.
    반환: {'寅': ['JC007', 'JC008', ...], '卯': [...], ..., '丑': [...]}
    """
    result = {}
    for branch in '寅卯辰巳午未申酉戌亥子丑':
        scored = [(_score_branch_job(branch, j), j['id']) for j in jobs]
        scored = [(s, code) for s, code in scored if s >= threshold]
        scored.sort(key=lambda x: (-x[0], x[1]))
        result[branch] = [code for _, code in scored]
    return result


# ─────────────────────────────────────────────────────────────
# 시트20 자동 생성 (월지별 매칭 직업군)
# ─────────────────────────────────────────────────────────────
def refresh_sheet20(xlsx, branch_to_jobs):
    """
    시트20_월지별직업군매칭 시트를 자동 생성/갱신.
    시트19와 동일한 구조: 번호 / 월지 / 매칭 수 / 매칭 코드.
    """
    try:
        wb = openpyxl.load_workbook(xlsx)
    except Exception as e:
        print(f'  WARN: 엑셀 로드 실패 ({e}). 시트20 스킵.')
        return

    # 시트 없으면 새로 생성 (+ 간단한 서식)
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    SHEET_NAME = '시트20_월지별직업군매칭'

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    title_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')

    # 대제목
    ws.merge_cells('A1:D1')
    ws['A1'] = '━━━ 12월지 × 매칭 직업군 코드 (시트18 기반 자동 생성) ━━━'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(vertical='center', horizontal='center')
    ws.row_dimensions[1].height = 34

    # 안내문
    ws.merge_cells('A2:D2')
    ws['A2'] = ('※ 이 시트는 excel_to_json.py 실행 시 시트18의 오행/십성/특수격/기질 태그와 '
                '시트17의 월지 오행·계절·본기장간을 참조하여 자동 생성됩니다. 월지 편집 없이 '
                '시트18만 수정하면 이 시트와 saju_db.json이 함께 갱신됩니다.')
    ws['A2'].font = Font(italic=True, size=10, color='555555')
    ws['A2'].fill = PatternFill(start_color='F5F5E9', end_color='F5F5E9', fill_type='solid')
    ws['A2'].alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
    ws.row_dimensions[2].height = 36

    # 헤더
    headers = ['번호', '월지', '매칭 직업군 수', '매칭 직업군 코드']
    for c, h in enumerate(headers, 1):
        cell_ = ws.cell(row=3, column=c)
        cell_.value = h
        cell_.font = Font(bold=True, color='FFFFFF', size=11)
        cell_.fill = header_fill
        cell_.alignment = Alignment(vertical='center', horizontal='center')
        cell_.border = thin
    ws.row_dimensions[3].height = 34

    # 월지별 오행 배경색
    BRANCH_FILL = {
        '寅':'E2EFDA', '卯':'C6E0B4',   # 木
        '巳':'FCE4D6', '午':'F8CBAD',   # 火
        '辰':'FFF2CC', '未':'FFE699', '戌':'FFEFB0', '丑':'FBE5C4',  # 土
        '申':'EDEDED', '酉':'D9D9D9',   # 金
        '亥':'DEEBF7', '子':'BDD7EE',   # 水
    }

    # 12월지 (寅卯辰...丑)
    order = list('寅卯辰巳午未申酉戌亥子丑')
    for idx, br in enumerate(order):
        r = 4 + idx
        codes = branch_to_jobs.get(br, [])
        ws.cell(row=r, column=1).value = idx + 1
        ws.cell(row=r, column=2).value = br
        ws.cell(row=r, column=3).value = len(codes)
        ws.cell(row=r, column=4).value = ', '.join(codes) if codes else ''

        bg = PatternFill(start_color=BRANCH_FILL.get(br, 'FFFFFF'),
                         end_color=BRANCH_FILL.get(br, 'FFFFFF'), fill_type='solid')
        for c in range(1, 5):
            cell_ = ws.cell(row=r, column=c)
            cell_.fill = bg
            cell_.border = thin
            if c in (1, 2, 3):
                cell_.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            else:
                cell_.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
            if c == 2:
                cell_.font = Font(bold=True, size=12)
        ws.row_dimensions[r].height = 72

    # 열 너비
    for col, w in {'A': 6, 'B': 10, 'C': 14, 'D': 110}.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A4'

    try:
        wb.save(xlsx)
    except Exception as e:
        print(f'  WARN: 시트20 엑셀 쓰기 실패 ({e}). JSON은 정상 생성됨.')


# ─────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────
def main(xlsx=DEFAULT_XLSX, json_out=DEFAULT_JSON):
    print(f'Reading {xlsx}...')

    if not os.path.exists(xlsx):
        print(f'ERROR: {xlsx} 파일이 없습니다.')
        sys.exit(1)

    # 모든 시트 추출
    month, harmony = extract_month_correction(xlsx)
    jobs_100, jc000 = extract_jobs_100(xlsx)
    ilju_to_jobs = refresh_sheet19(xlsx, jobs_100)
    branch_to_jobs = build_branch_to_jobs(jobs_100)
    refresh_sheet20(xlsx, branch_to_jobs)

    result = {
        'dna':                 extract_dna(xlsx),
        'gender_traits':       extract_gender(xlsx),
        'ilju_unique':         extract_ilju(xlsx),
        'sipsung_jobs':        extract_sipsung_jobs(xlsx),
        'job_visual':          extract_job_visual(xlsx),
        'body_pct':            extract_body_pct(xlsx),
        'ilju_combo':          extract_ilju_combo(xlsx),
        'month_correction':    month,
        'branch_harmony':      harmony,
        'job_categories_100':  jobs_100,
        'job_exclude_option':  jc000,
        'ilju_to_jobs':        ilju_to_jobs,
        'branch_to_jobs':      branch_to_jobs,
        'meta': {
            'schema_version': '2.1',
            'generated_at':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source':         os.path.basename(xlsx),
        },
    }

    with open(json_out, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, separators=(',', ':'))

    sz = os.path.getsize(json_out) / 1024
    print(f'Done -> {json_out} ({sz:.0f}KB)')
    print(f'  DNA:           {len(result["dna"])}')
    print(f'  Gender:        {len(result["gender_traits"])}')
    print(f'  Ilju:          {len(result["ilju_unique"])}')
    print(f'  SipsungJobs:   {len(result["sipsung_jobs"])}')
    print(f'  JobVisual:     {len(result["job_visual"])}')
    print(f'  BodyPct:       {len(result["body_pct"])}')
    print(f'  IljuCombo:     {len(result["ilju_combo"])}')
    print(f'  MonthCorr:     {len(result["month_correction"])}')
    print(f'  Harmony:       {len(result["branch_harmony"])}')
    print(f'  Jobs100:       {len(result["job_categories_100"])}')
    print(f'  IljuToJobs:    {len(result["ilju_to_jobs"])} iljus indexed')
    print(f'  BranchToJobs:  {len(result["branch_to_jobs"])} branches indexed')


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out  = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_JSON
    main(xlsx, out)
